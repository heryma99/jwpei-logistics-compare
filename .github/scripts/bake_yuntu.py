# -*- coding: utf-8 -*-
# Bake 云途(Yuntu) 小包渠道 from 云途/latest.xlsx (2026-8-4 生效).
# 仅更新已存在渠道的 brackets；保留原 countries 集合与 min_charge_weight/charge_rule 等元数据。
# 不确定来源的渠道(产品代码不在 2026-8-4 报价主表)默认不更新, 仅报告。
# Dry-run by default; --write to apply into rates.json.
import openpyxl, json, re, os, argparse

HERE = os.path.dirname(os.path.abspath(__file__))
NEWXLSX = os.path.join(HERE, "云途", "latest.xlsx")
RATES = os.path.join(HERE, "..", "..", "rates.json")
OUT = os.path.join(HERE, "patch_yuntu.json")

# 有确定来源的渠道 -> sheet
CONFIRMED = {
    "yun_thphr":       "云途全球专线挂号（特惠普货）",
    "yun_thp_battery": "云途全球专线挂号（特惠带电）",
}

def num(x):
    try:
        if x is None: return None
        s = str(x).replace("，", "").replace(",", "").strip()
        if s in ("", "***", "/", "-", "—"): return None
        return float(s)
    except: return None

def extract_brackets(sheet):
    """return {country: [{up, rate, reg}, ...]}"""
    wb = openpyxl.load_workbook(NEWXLSX, read_only=True, data_only=True)
    ws = wb[sheet]
    hidx = widx = ridx = gidx = None
    for ri in range(1, 14):
        row = [c for c in next(ws.iter_rows(min_row=ri, max_row=ri, max_col=14, values_only=True))]
        for i, v in enumerate(row):
            s = str(v) if v is not None else ""
            if s.startswith("国家/地区"): hidx = i
            elif "重量" in s and "KG" in s: widx = i
            elif "运费" in s and "KG" in s: ridx = i
            elif "挂号" in s: gidx = i
        if hidx is not None and widx is not None and ridx is not None and gidx is not None:
            break
    if hidx is None:
        wb.close(); raise RuntimeError(f"未找到表头于 {sheet}")
    brackets = {}
    last_country = None
    for row in ws.iter_rows(min_row=ri + 1, values_only=True):
        cells = list(row)
        if len(cells) <= max(hidx, widx, ridx, gidx):
            continue
        country = cells[hidx]
        if country and str(country).strip():
            last_country = str(country).strip()
        if last_country is None:
            continue
        wtext = cells[widx]; rate = cells[ridx]; reg = cells[gidx]
        if wtext is None or rate is None:
            continue
        m = re.search(r"≤\s*([\d.]+)", str(wtext))
        if not m:
            continue
        try:
            up = float(m.group(1)); rate = float(rate)
            reg = float(reg) if reg not in (None, "") else 0.0
        except Exception:
            continue
        brackets.setdefault(last_country, []).append({"up": up, "rate": rate, "reg": reg})
    wb.close()
    return brackets

def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--write", action="store_true"); args = ap.parse_args()
    rates = json.load(open(RATES, encoding="utf-8"))
    byid = {c["id"]: c for c in rates["channels"]}

    patch = {}
    report = []
    for cid, sheet in CONFIRMED.items():
        ch = byid.get(cid)
        if not ch:
            report.append(f"[跳过] {cid}: rates.json 无此渠道"); continue
        try:
            new = extract_brackets(sheet)
        except Exception as e:
            report.append(f"[错误] {cid}: {e}"); continue
        old = ch.get("countries", {})
        new_countries = {}
        updated = []; kept_old = []
        for country, old_data in old.items():
            if country in new:
                extra = {k: v for k, v in old_data.items() if k != "brackets"}
                new_countries[country] = {**extra, "brackets": new[country]}
                updated.append(country)
            else:
                new_countries[country] = old_data
                kept_old.append(country)
        ch_dest = list(new_countries.keys())
        patch[cid] = {"countries": new_countries, "dest": ch_dest}
        report.append(f"[更新] {cid} <- {sheet}: 更新国 {len(updated)} | 保留旧国(新报价无) {sorted(kept_old)}")
        us_old = old.get("美国", {}).get("brackets", [])
        us_new = new.get("美国", [])
        if us_old and us_new:
            report.append(f"        美国 旧首档 {us_old[0]} -> 新首档 {us_new[0]} | 档数 {len(us_old)}->{len(us_new)}")

    # 不确定性渠道(产品代码不在 2026-8-4 主表)报告, 不写盘
    ambiguous = {
        "yun_region":    "QYTHPH 区域专线特惠普货 (主表无此产品)",
        "yun_global_sp": "YTSPPH 全球商派标快普货 (主表无此产品)",
        "yun_ca_sp":     "YTCASPPH 加拿大商派标快普货 (主表无此产品)",
        "yun_a01":       "A01PH A01特惠普货 (主表无此产品)",
    }
    for cid, note in ambiguous.items():
        report.append(f"[待定] {cid}: {note} — 未更新, 需确认映射")

    print("\n".join(report))
    json.dump(patch, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\nWROTE {OUT} channels: {len(patch)}")
    if not args.write:
        print("[DRY-RUN] re-run with --write to apply")
        return
    for cid, p in patch.items():
        ch = byid[cid]
        ch["countries"] = p["countries"]
        ch["dest"] = p["dest"]
    json.dump(rates, open(RATES, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\n[WRITE] applied {len(patch)} channels")

if __name__ == "__main__":
    main()
