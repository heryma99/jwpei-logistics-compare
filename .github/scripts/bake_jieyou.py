# -*- coding: utf-8 -*-
# Bake 捷邮(jieyou) 中东专线小包 from 捷邮/latest.xlsx (如「捷邮中东专线小包报价表3.30.xlsx」).
# 仅更新已存在渠道 jieyou_mideast 的 countries（按目的国匹配）；保留结构(单段 up=30)与元数据。
# 不处理 jieyou_epost（国际E邮宝，报价单未覆盖，属手工维护）。
# Dry-run by default; --write to apply into rates.json.
import openpyxl, json, re, os, argparse

HERE = os.path.dirname(os.path.abspath(__file__))
NEWXLSX = os.path.join(HERE, "捷邮", "latest.xlsx")
RATES = os.path.join(HERE, "..", "..", "rates.json")
OUT = os.path.join(HERE, "patch_jieyou.json")

CHANNEL_ID = "jieyou_mideast"
# 报价单目的国简称 -> rates.json 国家全名
NAME_MAP = {
    "沙特": "沙特阿拉伯",
    "阿联酋": "阿拉伯联合酋长国",
    "巴林": "巴林",
    "卡塔尔": "卡塔尔",
    "科威特": "科威特",
    "阿曼": "阿曼",
}

def num(x):
    try:
        if x is None:
            return None
        s = str(x).replace(",", "").replace("，", "").strip()
        if s in ("", "/", "-", "—", "***"):
            return None
        return float(s)
    except Exception:
        return None

def parse_vat(x):
    if x is None:
        return 0.0
    m = re.search(r"(\d+(?:\.\d+)?)\s*%", str(x))
    return float(m.group(1)) / 100.0 if m else 0.0

def find_sheet(wb):
    for ws in wb.worksheets:
        for r in range(1, 12):
            row = [c for c in next(ws.iter_rows(min_row=r, max_row=r, max_col=10, values_only=True))]
            joined = " ".join(str(v) for v in row if v is not None)
            if "目的国" in joined and "运价" in joined:
                return ws, r
    return None, None

def locate_cols(ws):
    # 扫描前 12 行所有列，累积每列文本后定位（表头可能跨两行，如第6行战争附加费/第7行目的国）
    texts = {}
    for ri in range(1, 13):
        row = [c for c in next(ws.iter_rows(min_row=ri, max_row=ri, max_col=10, values_only=True))]
        for idx, v in enumerate(row):
            if v is not None:
                texts[idx] = texts.get(idx, "") + " " + str(v)
    def find(*kws):
        for idx, t in texts.items():
            if all(k in t for k in kws):
                return idx
        return None
    dest = find("目的国")
    rate = find("运价")
    op = find("操作费") or find("派送费")
    war = find("战争附加费")
    cust = find("海关资料处理费")
    vat = find("VAT") or find("DUTY")
    return dest, rate, op, war, cust, vat

def extract():
    wb = openpyxl.load_workbook(NEWXLSX, read_only=True, data_only=True)
    ws, hstart = find_sheet(wb)
    if ws is None:
        wb.close()
        raise RuntimeError("未找到含『目的国/运价』表头的 sheet")
    dest_i, rate_i, op_i, war_i, cust_i, vat_i = locate_cols(ws)
    if None in (dest_i, rate_i, op_i, war_i, cust_i, vat_i):
        wb.close()
        raise RuntimeError(f"列定位失败 dest={dest_i} rate={rate_i} op={op_i} war={war_i} cust={cust_i} vat={vat_i}")
    # 动态定位数据起始行：从 hstart+1 起，dest 列首个非空行
    data_start = None
    for ri in range(hstart + 1, ws.max_row + 1):
        rowvals = [c for c in next(ws.iter_rows(min_row=ri, max_row=ri, max_col=10, values_only=True))]
        if dest_i < len(rowvals) and rowvals[dest_i] is not None:
            data_start = ri
            break
    if data_start is None:
        wb.close()
        raise RuntimeError("未找到目的国数据行")
    out = {}
    for ri in range(data_start, ws.max_row + 1):
        cells = [c for c in next(ws.iter_rows(min_row=ri, max_row=ri, max_col=10, values_only=True))]
        if dest_i >= len(cells):
            continue
        dest = cells[dest_i]
        if dest is None:
            continue
        dest = str(dest).strip()
        rate = num(cells[rate_i]) if rate_i < len(cells) else None
        op = num(cells[op_i]) if op_i < len(cells) else None
        war = num(cells[war_i]) if war_i < len(cells) else None
        cust = num(cells[cust_i]) if cust_i < len(cells) else None
        vat = parse_vat(cells[vat_i]) if vat_i < len(cells) else 0.0
        if rate is None and op is None:
            continue
        out[dest] = {"rate": rate or 0.0, "op": op or 0.0, "war": war or 0.0, "cust": cust or 0.0, "vat": vat}
    wb.close()
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()
    if not os.path.exists(NEWXLSX):
        print(f"[跳过] 无 {NEWXLSX}（尚未抓到捷邮报价单）")
        return
    rates = json.load(open(RATES, encoding="utf-8"))
    byid = {c["id"]: c for c in rates["channels"]}
    ch = byid.get(CHANNEL_ID)
    if not ch:
        print(f"[跳过] {CHANNEL_ID}: rates.json 无此渠道")
        return
    try:
        latest = extract()
    except Exception as e:
        print(f"[错误] {CHANNEL_ID}: {e}")
        return
    co = ch.get("countries", {})
    patch = {}
    for dest, v in latest.items():
        rn = NAME_MAP.get(dest, dest)
        if rn not in co:
            print(f"  [跳过] 未匹配国家: {dest} -> {rn}")
            continue
        cur = co[rn]
        b = cur["brackets"][0]
        b["rate"] = v["rate"]
        b["reg"] = v["op"]
        cur.setdefault("add", {})["war_per_kg"] = v["war"]
        cur["add"]["custom_per_shipment"] = v["cust"]
        cur["tax"] = v["vat"]
        patch[rn] = {"rate": v["rate"], "reg": v["op"], "war": v["war"], "cust": v["cust"], "vat": v["vat"]}
    report = [f"[更新] {CHANNEL_ID} <- 捷邮中东专线小包报价表: 匹配国家 {len(patch)}/{len(latest)}"]
    for k, v in patch.items():
        report.append(f"    {k}: rate={v['rate']} reg={v['reg']} war={v['war']} cust={v['cust']} vat={v['vat']}")
    print("\n".join(report))
    json.dump({CHANNEL_ID: patch}, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\nWROTE {OUT}")
    if not args.write:
        print("[DRY-RUN] re-run with --write to apply")
        return
    json.dump(rates, open(RATES, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\n[WRITE] applied {CHANNEL_ID}")

if __name__ == "__main__":
    main()
