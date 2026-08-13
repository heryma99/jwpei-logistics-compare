# -*- coding: utf-8 -*-
"""中运通达商业快递 5 渠道 烘焙 (默认 dry-run; 加 --write 才写文件)"""
import json, re, sys, os
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
RATES = os.path.join(BASE, "..", "..", "rates.json")
XLSX = os.path.join(BASE, "中运通达", "latest.xlsx")
DRY = "--write" not in sys.argv

data = json.load(open(RATES, encoding="utf-8"))
chs = {c["id"]: c for c in data["channels"]}
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

def get_sheet(name):
    for sn in wb.sheetnames:
        if name in sn:
            return wb[sn]
    return None
def rows_of(ws): return list(ws.iter_rows(values_only=True))
def num(v): return isinstance(v,(int,float)) and v is not None

# ---- 燃油 ----
fuel_map = {}
cat = get_sheet("目录")
if cat:
    for r in rows_of(cat):
        for cell in r:
            if isinstance(cell,str) and "周燃油" in cell:
                for t in re.findall(r'(DHL|UPS|FEDEX)\D*?([\d.]+)%', cell, re.I):
                    fuel_map[t[0].upper()] = float(t[1])/100
print("燃油:", fuel_map)

def wkey(w):
    wf = float(w)
    return str(int(wf)) if wf==int(wf) else str(wf)

# ===== DHL (zy_dhl) =====
def bake_dhl():
    ch = chs["zy_dhl"]; ws=get_sheet("广州DHL"); rs=rows_of(ws)
    hi=next(i for i,r in enumerate(rs) if r[1]=="重量Weight (kg)")
    zones=[v for v in rs[hi][2:] if v not in (None,"返回目录")]; nz=len(zones)
    cr=next(i for i,r in enumerate(rs) if r[1]=="文件")
    pr=next(i for i,r in enumerate(rs) if r[1]=="包裹")
    br=next(i for i,r in enumerate(rs) if r[1]=="区间")
    countries=[rs[hi+1][2+i] for i in range(nz)]
    mdoc,m={},{}
    for ri in range(cr+1,pr):
        w=rs[ri][1]
        if num(w): mdoc[str(w)]=[float(rs[ri][2+i]) for i in range(nz)]
    for ri in range(pr+1,br):
        w=rs[ri][1]
        if num(w): m[str(w)]=[float(rs[ri][2+i]) for i in range(nz)]
    bt=[]
    for ri in range(br+1,br+6):
        lab=rs[ri][1]; mm=re.match(r'(\d+)-(\d+)',str(lab))
        if mm: bt.append({"min":int(mm.group(1)),"max":int(mm.group(2)),"unit":"kg",
                          "rates":[float(rs[ri][2+i]) for i in range(nz)]})
    ks=sorted(m,key=float)
    print(f"\n[DHL] zones={nz} matrix权重={len(ks)}({ks[0]}~{ks[-1]}) doc={len(mdoc)} bulk={len(bt)}")
    print(f"  matrix['0.5']={m['0.5'][:3]} matrix['20']={m['20'][:3]} doc['0.5']={mdoc.get('0.5',[])[:3]}")
    if not DRY:
        ch["zones"]=[{"label":zones[i],"countries":countries[i]} for i in range(nz)]
        ch["matrix"]=m; ch["matrix_doc"]=mdoc; ch["bulk_tiers"]=bt
        ch["fuel"]=fuel_map["DHL"]; ch["fuel_included"]=False
        print("  >> 写入 zy_dhl")

# ===== FEDEX (zy_fedex) =====
def bake_fedex():
    ch=chs["zy_fedex"]; ws=get_sheet("广州联邦IP"); rs=rows_of(ws)
    hi=next(i for i,r in enumerate(rs) if r[1]=="分区" and num(r[2]))
    codes=[rs[hi][2+i] for i in range(20)]
    cnames=[rs[hi+1][2+i] for i in range(20)]
    nz=20
    pkg_row=next(i for i,r in enumerate(rs) if r[1]=="包裹")
    mdoc={}
    for ri in range(hi,len(rs)):
        if rs[ri][1]=="文件0.5":
            mdoc["0.5"]=[float(rs[ri][2+i]) for i in range(nz)]
    mpak={}
    for ri in range(hi+1,pkg_row):
        w=rs[ri][1]
        if num(w): mpak[str(w)]=[float(rs[ri][2+i]) for i in range(nz)]
    m={}
    for ri in range(pkg_row+1,len(rs)):
        w=rs[ri][1]
        if num(w): m[str(w)]=[float(rs[ri][2+i]) for i in range(nz)]
    ks=sorted(m,key=float)
    print(f"\n[FEDEX] zones={nz} matrix权重={len(ks)}({ks[0]}~{ks[-1]}) pak={len(mpak)} doc={len(mdoc)}")
    print(f"  matrix['0.5']={m['0.5'][:3]} pak['0.5']={mpak.get('0.5',[])[:3]} doc={list(mdoc.items())[:1]}")
    if not DRY:
        ch["zones"]=[{"label":str(codes[i]),"countries":cnames[i]} for i in range(nz)]
        ch["matrix"]=m; ch["matrix_doc"]=mdoc; ch["matrix_pak"]=mpak
        ch["fuel"]=fuel_map["FEDEX"]; ch["fuel_included"]=False
        print("  >> 写入 zy_fedex")

# ===== UPS小货 (zy_ups) =====
def bake_ups():
    ch=chs["zy_ups"]; ws=get_sheet("大陆UPS-红单小货"); rs=rows_of(ws)
    hi=next(i for i,r in enumerate(rs) if r[1]=="分区" and isinstance(r[2],str))
    zones=[v for v in rs[hi][2:] if v not in (None,"返回目录")]; nz=len(zones)
    m={}
    for ri in range(hi+2,len(rs)):
        w=rs[ri][1]
        if num(w):
            vals=[rs[ri][2+i] for i in range(nz)]
            if all(num(v) for v in vals): m[str(w)]=[float(v) for v in vals]
    ks=sorted(m,key=float)
    print(f"\n[UPS小货] zones={nz} matrix权重={len(ks)}({ks[0]}~{ks[-1]})")
    print(f"  matrix['0.5']={m['0.5'][:3]} matrix['20']={m.get('20',[])[:3]}")
    if not DRY:
        ch["zones"]=[{"label":zones[i],"countries":zones[i]} for i in range(nz)]
        ch["matrix"]=m
        print("  >> 写入 zy_ups (countryMap/sur 保留)")

# ===== UPS大货 (zy_ups_bulk) =====
def bake_ups_bulk():
    ch=chs["zy_ups_bulk"]; ws=get_sheet("大陆UPS-红单大货"); rs=rows_of(ws)
    rates=[]
    for ri in (3,4,5):
        nums=[v for v in rs[ri][2:] if num(v)]
        if nums: rates.append(float(nums[0]))
    print(f"\n[UPS大货] 标准价 美/加墨/欧 = {rates}")
    if not DRY:
        for seg in ch["bulk_tiers"]: seg["rates"]=rates
        print("  >> 写入 zy_ups_bulk bulk_tiers")

# ===== UPS包税美国 (zy_ups_us) =====
def bake_ups_us():
    ch=chs["zy_ups_us"]; ws=get_sheet("大陆UPS包税(美国)"); rs=rows_of(ws)
    m={}
    for r in rs:
        cells=list(r); i=1
        while i+2 < len(cells):
            wl,ca,us=cells[i],cells[i+1],cells[i+2]
            if isinstance(wl,str) and num(ca) and num(us):
                mm=re.match(r'([\d.]+)',wl)
                if mm: m[wkey(mm.group(1))]=[float(ca),float(us)]
            i+=5
    ks=sorted(m,key=float)
    print(f"\n[UPS包税美国] matrix权重={len(ks)}({ks[0]}~{ks[-1]})")
    print(f"  matrix['0.5']={m.get('0.5')} matrix['21']={m.get('21')}")
    if not DRY:
        ch["matrix"]=m
        print("  >> 写入 zy_ups_us")

for fn in (bake_dhl,bake_fedex,bake_ups,bake_ups_bulk,bake_ups_us):
    fn()

if not DRY:
    json.dump(data, open(RATES,"w",encoding="utf-8"), ensure_ascii=False, indent=1)
    print("\n*** 已写入 rates.json ***")
else:
    print("\n[dry-run] 未写入。加 --write 执行真实写入。")
wb.close()
