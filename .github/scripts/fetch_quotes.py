# -*- coding: utf-8 -*-
"""
fetch_quotes.py — 云端定时抓报价单并自动更新基础价格（GitHub Actions 内运行）。
流程: 用 wl02@ 的 user_access_token 搜邮箱报价单 -> 下附件 -> 解压选主表
      -> 调各 bake_*.py 解析(只改价格数值) -> 白名单 diff 守铁律
      -> 若有变化: 重新生成 rates.js + 飞书群(比价易通知)推**环比变动快照图**（仅图片，不发文字/卡片）。
依赖: openpyxl（bake 需要）；user token 来自 secret WL02_MAIL_TOKEN（JSON）。
本地回归: 设 FETCH_LOCAL=1 则从本地 quote_pull/{carrier}/latest.xlsx 复制，不联网。
"""
import json, os, sys, uuid, urllib.request, urllib.parse, zipfile, shutil, subprocess, time, datetime, re

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(os.path.dirname(HERE))          # .github/scripts -> repo root
RATES = os.path.join(REPO_ROOT, "rates.json")
RATES_JS = os.path.join(REPO_ROOT, "rates.js")
VERSION = os.path.join(REPO_ROOT, "version.json")

APP_ID = "cli_aab6341b78f95be9"
APP_SECRET = os.environ.get("FEISHU_APP_SECRET", "")   # 云端经 workflow secrets 注入，禁止硬编码
MBOX = "wl02@leading-trade.com"
CHAT_ID = "oc_cd9a6f072cdd348a08c29d09e8c9143a"
SITE_URL = "https://heryma99.github.io/jwpei-logistics-compare/"

CARRIERS = {
    "云途":     {"kw": ["云途", "客户报价单"], "main": "云途全球专线挂号（特惠普货）", "prefer_ext": "zip"},
    "中运通达": {"kw": ["中运通达", "中运"],    "main": None, "prefer_ext": "xlsx"},
    "亚丰":     {"kw": ["璞景", "德翼供应链", "德翼价格表", "德翼", "亚丰"], "main": None, "prefer_ext": "xlsx"},
    "飞特":     {"kw": ["飞特", "飞特标准挂号", "飞特物流"], "main": None, "prefer_ext": "xlsx"},
}
QUOTE_SOURCES = {}   # 抓取时记录各家选定报价邮件主题，供 bake 备注真实报价日
BAKE = {
    "云途":     ["bake_yuntu.py"],
    "中运通达": ["bake_zy_package.py", "bake_commercial.py"],
    "亚丰":     ["bake_yf.py"],
    "飞特":     ["bake_feit.py"],
}
MAIL_API = "https://open.feishu.cn/open-apis/mail/v1/user_mailboxes/" + MBOX
REFRESH_URL = "https://open.feishu.cn/open-apis/authen/v1/oidc/refresh_access_token"

FETCH_LOCAL = os.environ.get("FETCH_LOCAL", "") == "1"
LOCAL_XLSX_DIR = os.environ.get("LOCAL_XLSX_DIR", os.path.join(REPO_ROOT, "quote_pull"))

# 每趟运行对各承运商的处理结果汇总（用于诊断卡片 / 日志）
REPORT = []


def log(m):
    print(f"[fetch] {m}", flush=True)


def req_json(url, method="GET", token=None, body=None):
    h = {"Content-Type": "application/json", "User-Agent": "fetch-quotes"}
    if token:
        h["Authorization"] = "Bearer " + token
    r = urllib.request.Request(url, data=json.dumps(body).encode() if body is not None else None,
                               method=method, headers=h)
    with urllib.request.urlopen(r, timeout=60) as resp:
        return json.loads(resp.read().decode("utf-8"))


# ---------- token ----------
def get_app_access_token():
    """用 App Secret 换取 app_access_token，供 authen 系列接口鉴权。"""
    if not APP_SECRET:
        raise RuntimeError("FEISHU_APP_SECRET 未配置")
    url = "https://open.feishu.cn/open-apis/auth/v3/app_access_token/internal"
    r = req_json(url, "POST", body={"app_id": APP_ID, "app_secret": APP_SECRET})
    if r.get("code") != 0:
        raise RuntimeError(f"获取 app_access_token 失败: {r}")
    return r["app_access_token"]


def load_token():
    raw = os.environ.get("WL02_MAIL_TOKEN", "")
    if not raw:
        log("[skip] 无 WL02_MAIL_TOKEN secret，跳过本次抓取")
        return None
    try:
        tok = json.loads(raw)
    except Exception:
        log("[skip] WL02_MAIL_TOKEN 解析失败")
        return None
    # 不再信任 access_token 的过期时间：飞书 access_token 可能提前失效，
    # 因此每次运行都先用 refresh_token 换一个新的 access_token，确保可用。
    if tok.get("refresh_token"):
        try:
            app_token = get_app_access_token()
            r = req_json(REFRESH_URL, "POST", token=app_token, body={
                "grant_type": "refresh_token",
                "refresh_token": tok["refresh_token"],
            })
            if r.get("code") == 0:
                data = r.get("data", r)
                tok["access_token"] = data["access_token"]
                tok["refresh_token"] = data.get("refresh_token", tok["refresh_token"])
                tok["expires_in"] = data.get("expires_in", 7200)
                tok["got_at"] = int(time.time())
                log("refresh 成功")
                save_token_to_secret(tok)
                return tok["access_token"]
            else:
                log(f"refresh 失败: {r}")
        except Exception as e:
            log(f"refresh 失败: {e}")
    # refresh 失败时，如果仍有一个 access_token 就先拿来用（大概率也会失败，但至少能试）
    if tok.get("access_token"):
        log("[warn] refresh 失败，回退使用现有 access_token")
        return tok["access_token"]
    log("[skip] 无可用 access_token")
    return None


def save_token_to_secret(tok):
    """刷新成功后把新 token 写回 GitHub secret，实现零维护自动续期。
    需 workflow 提供 GH_PAT（Personal Access Token，含 repo scope）才能写 secrets；
    GITHUB_TOKEN 默认无写 secret 权限，故仅作为 fallback。失败仅告警不影响本次抓取。"""
    gt = os.environ.get("GH_PAT") or os.environ.get("GITHUB_TOKEN")
    repo = os.environ.get("GITHUB_REPO")
    if not gt or not repo:
        log("[skip] 未配置 GITHUB_TOKEN/GITHUB_REPO，跳过写回 secret（后续过期需手动重授权）")
        return
    try:
        import base64 as _b64
        from nacl.public import PublicKey, SealedBox
        import nacl.encoding
        H = {"Authorization": f"Bearer {gt}", "Accept": "application/vnd.github+json",
             "Content-Type": "application/json", "X-GitHub-Api-Version": "2022-11-28"}
        def _api(method, path, body=None):
            data = json.dumps(body).encode() if body is not None else None
            req = urllib.request.Request("https://api.github.com" + path, data=data, headers=H, method=method)
            with urllib.request.urlopen(req, timeout=30) as r:
                b = r.read().decode()
                return r.status, (json.loads(b) if b.strip() else {})
        st, keyinfo = _api("GET", f"/repos/{repo}/actions/secrets/public-key")
        if st != 200:
            log(f"[warn] 获取 public-key 失败 {st}"); return
        pk = PublicKey(keyinfo["key"].encode(), encoder=nacl.encoding.Base64Encoder)
        box = SealedBox(pk)
        enc = nacl.encoding.Base64Encoder.encode(box.encrypt(json.dumps(tok).encode())).decode()
        st2, _ = _api("PUT", f"/repos/{repo}/actions/secrets/WL02_MAIL_TOKEN",
                      {"encrypted_value": enc, "key_id": keyinfo["key_id"]})
        if st2 in (201, 204):
            log("✅ 新 token 已自动写回 WL02_MAIL_TOKEN secret（零维护自动续期生效）")
        else:
            log(f"[warn] 写回 secret 返回 {st2}")
    except Exception as e:
        log(f"[warn] 写回 secret 失败（不影响本次抓取）: {e}")


# ---------- 邮件 ----------
def search_mail(token, kw):
    # 飞书邮件搜索正确端点：POST /open-apis/mail/v1/user_mailboxes/{MBOX}/search
    # （注意：不是 /messages/search，那个路径返回 404）
    try:
        r = req_json(MAIL_API + "/search", "POST", token, {"query": kw, "page_size": 15})
    except Exception as e:
        log(f"search '{kw}' 异常: {e}")
        REPORT.append(f"❌ 搜索「{kw}」异常: {e}")
        return []
    if r.get("code") != 0:
        log(f"search '{kw}' 错误: {r.get('code')} {r.get('msg')}")
        REPORT.append(f"❌ 搜索「{kw}」错误码 {r.get('code')}: {r.get('msg')}")
        return []
    msgs = r.get("data", {}).get("items", []) or []
    REPORT.append(f"· 搜索「{kw}」: 命中 {len(msgs)} 封")
    return msgs


def get_attachments(token, mid):
    r = req_json(MAIL_API + f"/messages/{mid}", "GET", token)
    if r.get("code") != 0:
        return []
    return r.get("data", {}).get("message", {}).get("attachments", []) or []


def download_url(token, mid, aid):
    # 飞书下载接口：单个 id 直接拼在 query 上（不要传 JSON 数组）；
    # 响应字段是 data.download_urls[]（不是 url_list）
    r = req_json(MAIL_API + f"/messages/{mid}/attachments/download_url?attachment_ids={aid}", "GET", token)
    if r.get("code") != 0:
        return None
    for d in r.get("data", {}).get("download_urls", []):
        if d.get("attachment_id") == aid and d.get("download_url"):
            return d["download_url"]
    return None


def pick_target(atts, prefer_ext):
    # 真实附件 attachment_type 为真（如 1）；内联图片 attachment_type 为 None
    real = [a for a in atts if a.get("attachment_type") and not str(a.get("filename", "")).lower().startswith("image")]
    if prefer_ext:
        t = next((a for a in real if str(a.get("filename", "")).lower().endswith(prefer_ext)), None)
        if t:
            return t
    for ext in (".zip", ".xlsx", ".docx"):
        t = next((a for a in real if str(a.get("filename", "")).lower().endswith(ext)), None)
        if t:
            return t
    return real[0] if real else None


def pick_main_xlsx(zippath, main_sheet):
    import openpyxl, io
    # 直接内存读取 zip 内 xlsx 成员，避免 Windows 解压中文文件名乱码导致路径找不到
    with zipfile.ZipFile(zippath) as z:
        names = [n for n in z.namelist() if n.lower().endswith(".xlsx")]
        chosen = None
        if main_sheet:
            best = -1
            for n in names:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(z.read(n)), read_only=True)
                    if main_sheet in wb.sheetnames and len(wb.sheetnames) > best:
                        chosen, best = n, len(wb.sheetnames)
                    wb.close()
                except Exception:
                    pass
        if not chosen and names:
            chosen = max(names, key=lambda n: z.getinfo(n).date_time)
        if not chosen:
            return None
        tmp = zippath + ".x"
        open(tmp, "wb").write(z.read(chosen))
        return tmp


def is_price_mail(subj, fname):
    """判断是否为报价单邮件：排除账单/发票/对账单等，且需含价格/报价字样。"""
    s = (subj + " " + (fname or "")).lower()
    if any(k in s for k in ("账单", "对账单", "发票", "invoice", "明细", "statement", "请款", "月结")):
        return False
    return any(k in s for k in ("价格", "报价", "vip", "价目", "价格表", "价表"))


def process_carrier(token, name, cfg):
    outdir = os.path.join(HERE, name)
    os.makedirs(outdir, exist_ok=True)
    final = os.path.join(outdir, "latest.xlsx")
    # 手动上传覆盖：若 manual/<name>/latest.xlsx 存在，优先用它（不走邮箱）。
    # 用途：邮箱没收到某承运商报价单时（如 DPEX/2090/顺丰/易连/捷邮 只发月结账单），
    # 由用户提供报价单 xlsx 放此路径，走与邮箱完全相同的烘焙流程。
    manual_path = os.path.join(REPO_ROOT, "manual", name, "latest.xlsx")
    if os.path.exists(manual_path):
        shutil.copy(manual_path, final)
        log(f"[{name}] 手动上传覆盖：{manual_path}")
        REPORT.append(f"✅ [{name}] 使用手动上传的报价单（{os.path.basename(manual_path)}）")
        return True
    if FETCH_LOCAL:
        src = os.path.join(LOCAL_XLSX_DIR, name, "latest.xlsx")
        if not os.path.exists(src):
            log(f"[{name}] 本地回归缺 {src}")
            return False
        shutil.copy(src, final)
        log(f"[{name}] 本地回归用 {src}")
        REPORT.append(f"· [{name}] 本地回归用 {src}")
        return True
    # 联网抓取：收集所有关键词命中邮件（去重），按收件时间(internal_date)取最新且含报价附件的一封
    candidates = []
    seen = set()
    for kw in cfg["kw"]:
        items = search_mail(token, kw)
        for em in items:
            mid = em.get("message_id") or em.get("id")
            if mid in seen:
                continue
            seen.add(mid)
            atts = get_attachments(token, mid)
            tgt = pick_target(atts, cfg["prefer_ext"])
            if not tgt:
                continue
            d = req_json(MAIL_API + f"/messages/{mid}", "GET", token)
            msg = (d.get("data") or {}).get("message", {}) if d.get("code") == 0 else {}
            subj = msg.get("subject") or em.get("subject") or ""
            internal = str(msg.get("internal_date") or "")
            candidates.append({
                "mid": mid, "tgt": tgt, "subj": subj,
                "internal": internal,
                "is_price": is_price_mail(subj, tgt.get("filename", "")),
            })
    if not candidates:
        log(f"[{name}] 候选邮件均无报价附件")
        REPORT.append(f"· [{name}] 候选邮件均无报价附件")
        return False
    # 优先取「报价单」类，其次退化为全部候选；都按收件时间倒序取最新
    price_cands = [c for c in candidates if c["is_price"]]
    pool = price_cands if price_cands else candidates
    pool.sort(key=lambda c: c["internal"], reverse=True)
    best = pool[0]
    mid, tgt, subj = best["mid"], best["tgt"], best["subj"]
    log(f"[{name}] 选定最新报价邮件: {subj[:40]} (internal={best['internal']})")
    QUOTE_SOURCES[name] = subj
    url = download_url(token, mid, tgt["id"])
    if not url:
        log(f"[{name}] 拿不到下载链接")
        REPORT.append(f"❌ [{name}] 邮件有附件但拿不到下载链接")
        return False
    raw = os.path.join(outdir, "raw_" + re.sub(r"[^\w.]", "_", str(tgt.get("filename", "x"))))
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    data = urllib.request.urlopen(req, timeout=180).read()
    open(raw, "wb").write(data)
    if raw.lower().endswith(".zip"):
        cx = pick_main_xlsx(raw, cfg["main"])
        if cx:
            shutil.copy(cx, final)
            if cx != raw:
                os.remove(cx)
        os.remove(raw)
    else:
        shutil.copy(raw, final)
        os.remove(raw)
    log(f"[{name}] 已落盘 {final} ({len(data)} bytes)")
    REPORT.append(f"✅ [{name}] 已下载报价单 {tgt.get('filename')} ({len(data)} bytes) 并落盘")
    return True


# ---------- 白名单 diff（守铁律：只允许数值变化） ----------
def is_num_keys(d):
    if not d:
        return False
    for k in d:
        try:
            float(k)
        except (TypeError, ValueError):
            return False
    return True


def norm_keys(d):
    return {float(k): k for k in d} if (d and is_num_keys(d)) else None


def struct_check(path, a, b, rep):
    if isinstance(a, dict) and isinstance(b, dict):
        na, nb = norm_keys(a), norm_keys(b)
        for k in a:
            present = (k in b) if na is None else (float(k) in nb)
            if not present:
                rep.append("FAIL 删除字段 %s.%s" % (path, k)); return False
        for k in b:
            present = (k in a) if na is None else (float(k) in na)
            if not present:
                if is_num_keys(a) or is_num_keys(b):
                    rep.append("FAIL 档位键变化 %s:+%s" % (path, k)); return False
                rep.append("WARN 新增字段 %s.%s" % (path, k))
        for k in a:
            kb = k if na is None else nb.get(float(k))
            if kb is not None and not struct_check(path + "." + k, a[k], b[kb], rep):
                return False
        return True
    if isinstance(a, list) and isinstance(b, list):
        if len(a) != len(b):
            rep.append("FAIL 数组长度 %s:%d->%d" % (path, len(a), len(b))); return False
        for i, (x, y) in enumerate(zip(a, b)):
            if not struct_check("%s[%d]" % (path, i), x, y, rep):
                return False
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) \
            and not isinstance(a, bool) and not isinstance(b, bool):
        return True
    if type(a) != type(b):
        rep.append("FAIL 类型变化 %s:%s->%s" % (path, type(a).__name__, type(b).__name__)); return False
    return True


def whitelist_ok(old, new):
    oc = {c["id"]: c for c in old["channels"]}
    nc = {c["id"]: c for c in new["channels"]}
    rep = []
    for cid in (set(oc) ^ set(nc)):
        rep.append("FAIL 渠道增删: " + cid)
    for cid in set(oc) & set(nc):
        struct_check("channels." + cid, oc[cid], nc[cid], rep)
    fails = [r for r in rep if r.startswith("FAIL")]
    for r in rep:
        log(r)
    return not fails


# ---------- 变更卡片（复用 notify 逻辑，内存 old/new 对比） ----------
def flatten(o, p=""):
    out = {}
    if isinstance(o, dict):
        for k, v in o.items():
            out.update(flatten(v, f"{p}{k}.")) if isinstance(v, (dict, list)) else out.update({p + k: v})
    elif isinstance(o, list):
        for i, v in enumerate(o):
            out.update(flatten(v, f"{p}[{i}].")) if isinstance(v, (dict, list)) else out.update({f"{p}[{i}]": v})
    return out


def humanize(p):
    m = re.match(r"country_overrides\.(.+?)\.volumetric_divisor$", p)
    if m: return f"{m.group(1)} 体积重除数"
    if p.endswith("volumetric_divisor"): return "体积重除数"
    m = re.match(r"countries\.(.+?)\.service_fee_pct$", p)
    if m: return f"{m.group(1)} 服务费比例"
    m = re.match(r"countries\.(.+?)\.brackets\.\[(\d+)\]\.rate$", p)
    if m: return f"{m.group(1)} 第{int(m.group(2))+1}段费率(元/kg)"
    m = re.match(r"countries\.(.+?)\.brackets\.\[(\d+)\]\.(.+)$", p)
    if m: return f"{m.group(1)} 第{int(m.group(2))+1}段 {m.group(3)}"
    m = re.match(r"countries\.(.+?)\.tiers\.\[(\d+)\]\.rate$", p)
    if m: return f"{m.group(1)} 第{int(m.group(2))+1}段费率(元/kg)"
    m = re.match(r"countries\.(.+?)\.tiers\.\[(\d+)\]\.(.+)$", p)
    if m: return f"{m.group(1)} 第{int(m.group(2))+1}段 {m.group(3)}"
    if "matrix" in p: return "分区矩阵 " + p.split(".")[-1]
    return p.split(".")[-1]


def fmt_val(p, v):
    if isinstance(v, float) and "pct" in p:
        return f"{v*100:.2f}%"
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


def compute_diff(old_d, new_d):
    oc = {c["id"]: c for c in old_d["channels"]}
    nc = {c["id"]: c for c in new_d["channels"]}
    added = [c for c in nc if c not in oc]
    removed = [c for c in oc if c not in nc]
    changed = []
    for cid in sorted(set(oc) & set(nc)):
        a, b = flatten(oc[cid]), flatten(nc[cid])
        dc = [(humanize(k), fmt_val(k, a.get(k)) if a.get(k) is not None else "—",
               fmt_val(k, b.get(k)) if b.get(k) is not None else "—")
              for k in sorted(set(a) | set(b)) if a.get(k) != b.get(k)]
        if dc:
            changed.append((cid, nc[cid].get("name", cid), dc))
    return added, removed, changed


def build_card(committer, ctime, added, removed, changed):
    lines = [f"**更新人**：{committer}", f"**更新时间**：{ctime}",
             f"**变动规模**：{len(changed)} 个渠道价格变动（新增 {len(added)} / 移除 {len(removed)}）", ""]
    if changed:
        lines.append("📝 **价格变动明细**（自动抓单解析）：")
        for cid, name, dc in changed[:18]:
            lines.append(f"\n**{name}**  `{cid}`")
            for label, ov, nv in dc[:8]:
                lines.append(f"  · {label}：{ov} → {nv}")
            if len(dc) > 8:
                lines.append(f"  · …另 {len(dc)-8} 处")
        if len(changed) > 18:
            lines.append(f"\n…共 {len(changed)} 个渠道变动")
    text = "\n".join(lines)
    return {"config": {"wide_screen_mode": True},
            "header": {"title": {"tag": "plain_text", "content": "比价易 · 报价单自动更新"},
                       "template": "blue"},
            "elements": [{"tag": "div", "text": {"tag": "lark_md", "content": text}},
                         {"tag": "hr"},
                         {"tag": "action", "actions": [{"tag": "button", "text": {"tag": "plain_text", "content": "查看比价易"},
                                                       "type": "primary", "url": SITE_URL}]}]}


# ---------- 环比通知（纯文本，可直接转发） ----------
def _fmt_w(x):
    """重量数值格式化：去掉无意义的尾随 0（0.10->0.1, 2.0->2, 30.0->30）。"""
    if x is None:
        return ""
    x = float(x)
    if x == int(x):
        return str(int(x))
    return ("%g" % x)


def extract_price_rows(ch):
    """归一为 {(dim, w, kind): rate}：rate 为运费/kg（元）。覆盖各渠道真实价格字段。
    重量段一律还原成「报价里的实际区间（用 > ≤ < 符号表示，不用括号）」：
    - 云途 brackets：countries.<国>.brackets[] {up, rate, reg} —— 相邻 up 拼接成
      「>{prev_up} ≤{up}kg」（首段下限为 0）
    - 中运/亚丰 tiers：countries.<国>.tiers[] {min, rate} —— 相邻 min 拼接成
      「>{min} <{next_min}kg」（末段「>{min}kg」）
    - 中运 大货 To B：channel.bulk_tiers[] {min, max, rates[区]} —— 直接用显式
      「>{min} ≤{max}kg」
    - 中运 express：matrix[重量段][区] + fuel —— 相邻键拼接成「>{prev} ≤{w}kg」
    其它结构无价格行则返回空，不影响。"""
    rows = {}
    ctry = ch.get("countries")
    if isinstance(ctry, dict):
        for country, v in ctry.items():
            if not isinstance(v, dict):
                continue
            # 云途 brackets：相邻 up 拼接完整区间
            arr = v.get("brackets")
            if isinstance(arr, list) and arr:
                for i, seg in enumerate(arr):
                    if not isinstance(seg, dict):
                        continue
                    up = seg.get("up")
                    if up is None:
                        wlabel = f"第{i+1}段"
                    else:
                        lo = arr[i - 1].get("up") if i > 0 else 0
                        wlabel = f">{_fmt_w(lo)} ≤{_fmt_w(up)}kg"
                    if seg.get("rate") is not None:
                        rows[(country, wlabel, "rate")] = float(seg["rate"])
                    if seg.get("reg") is not None:
                        rows[(country, "挂号费", "reg")] = float(seg["reg"])
            # 中运/亚丰 tiers：相邻 min 拼接完整区间
            arr = v.get("tiers")
            if isinstance(arr, list) and arr:
                n = len(arr)
                for i, seg in enumerate(arr):
                    if not isinstance(seg, dict):
                        continue
                    mn = seg.get("min")
                    if mn is None:
                        wlabel = f"第{i+1}段"
                    else:
                        hi = arr[i + 1].get("min") if i + 1 < n else None
                        wlabel = (f">{_fmt_w(mn)} <{_fmt_w(hi)}kg"
                                  if hi is not None else f">{_fmt_w(mn)}kg")
                    if seg.get("rate") is not None:
                        rows[(country, wlabel, "rate")] = float(seg["rate"])
    # 中运 大货 To B：channel 级 bulk_tiers（显式 min/max，按区取价）
    arr = ch.get("bulk_tiers")
    if isinstance(arr, list) and arr:
        zones = ch.get("zones", [])
        zlabel = {i: (z.get("label", f"区{i+1}") if isinstance(z, dict) else f"区{i+1}") for i, z in enumerate(zones)}
        for seg in arr:
            if not isinstance(seg, dict):
                continue
            mn = seg.get("min"); mx = seg.get("max")
            if mn is None or mx is None:
                continue
            wlabel = f">{_fmt_w(mn)} ≤{_fmt_w(mx)}kg"
            rates = seg.get("rates")
            if isinstance(rates, list):
                for zi, rate in enumerate(rates):
                    if rate is None:
                        continue
                    rows[(zlabel.get(zi, f"区{zi+1}"), wlabel, "rate")] = float(rate)
    if "matrix" in ch and isinstance(ch["matrix"], dict):
        zones = ch.get("zones", [])
        zlabel = {i: (z.get("label", f"区{i+1}") if isinstance(z, dict) else f"区{i+1}") for i, z in enumerate(zones)}
        for w in sorted(ch["matrix"].keys(), key=lambda k: float(k)):
            arr = ch["matrix"][w]
            if not isinstance(arr, list):
                continue
            lower = 0.0
            # 找上一个键作下限
            prev = None
            for k in sorted(ch["matrix"].keys(), key=lambda k: float(k)):
                if k == w:
                    break
                prev = k
            if prev is not None:
                lower = float(prev)
            wlabel = f">{_fmt_w(lower)} ≤{_fmt_w(w)}kg"
            for zi, rate in enumerate(arr):
                if rate is None:
                    continue
                rows[(f"区{zlabel.get(zi, zi+1)}", wlabel, "rate")] = float(rate)
    if ch.get("fuel") is not None:
        rows[("(燃油率)", "整体", "fuel")] = float(ch["fuel"])
    return rows


def compute_ringbi(old_d, new_d):
    """返回 [(channel_dict, [change,...]), ...]，change={dim,w,kind,old,new,delta,pct,direction}。
    覆盖：新旧都有的渠道做逐项环比；仅在 new 出现的渠道视为整渠道「新增」。"""
    oc = {c["id"]: c for c in old_d["channels"]}
    nc = {c["id"]: c for c in new_d["channels"]}
    out = []
    for cid in sorted(set(oc) | set(nc)):
        oc_c = oc.get(cid)
        nc_c = nc.get(cid)
        if oc_c and nc_c:
            pr = extract_price_rows(oc_c)
            nr = extract_price_rows(nc_c)
            if not (pr or nr):
                continue
            changes = []
            for key in sorted(set(pr) | set(nr), key=lambda k: (k[0], k[1])):
                o = pr.get(key); n = nr.get(key)
                if o == n:
                    continue
                if o is None:
                    direction = "新增"
                elif n is None:
                    direction = "移除"
                else:
                    delta = n - o
                    pct = (delta / o * 100) if o else 0.0
                    direction = "涨" if delta > 0 else ("跌" if delta < 0 else "平")
                changes.append({
                    "dim": key[0], "w": key[1], "kind": key[2],
                    "old": o, "new": n,
                    "delta": (None if (o is None or n is None) else n - o),
                    "pct": (None if (o is None or n is None or not o) else (n - o) / o * 100),
                    "direction": direction,
                })
            if changes:
                out.append((nc_c, changes))
        elif nc_c and not oc_c:
            # 整渠道新增：old 缺失，所有价格行视为新增
            nr = extract_price_rows(nc_c)
            if not nr:
                continue
            changes = [{
                "dim": k[0], "w": k[1], "kind": k[2],
                "old": None, "new": v, "delta": None, "pct": None, "direction": "新增",
            } for k, v in sorted(nr.items(), key=lambda kv: (kv[0][0], kv[0][1]))]
            if changes:
                out.append((nc_c, changes))
        # 仅 old 有的渠道（被移除）不推送
    return out


def _fmt_money(x):
    return "—" if x is None else f"{x:,.2f}"


def _pct_str(p):
    return "—" if p is None else f"{p:+.1f}%"


def render_ringbi_text(rep, prev_label, new_label):
    L = []
    L.append("📊 物流报价环比变动通知")
    L.append(f"环比基准：{prev_label}  →  {new_label}")
    L.append(f"覆盖承运商：云途 / 中运通达 / 亚丰（仅价格数值变动，红涨绿跌·运费/kg口径）")
    L.append("=" * 54)
    up = sum(1 for _, chs in rep for x in chs if x["direction"] == "涨")
    dn = sum(1 for _, chs in rep for x in chs if x["direction"] == "跌")
    newc = sum(1 for _, chs in rep for x in chs if x["direction"] == "新增")
    maxjump = None
    for _, chs in rep:
        for x in chs:
            if x["pct"] is not None and (maxjump is None or abs(x["pct"]) > abs(maxjump[0])):
                maxjump = (x["pct"], x)
    L.append(f"【汇总】涨价 {up} ｜ 降价 {dn} ｜ 新增 {newc} ｜ 涉及渠道 {len(rep)}"
             + (f" ｜ 最大单跳：{_pct_str(maxjump[0])}" if maxjump else ""))
    L.append("=" * 54)
    for c, changes in rep:
        car = c.get("carrier", "") or c.get("name", "")
        L.append(f"\n▌{c.get('name')}  （{car}）")
        for x in changes:
            arrow = {"涨": "🔺", "跌": "🟢", "新增": "🆕", "移除": "➖", "平": "➖"}.get(x["direction"], "")
            if x["kind"] == "fuel":
                L.append(f"   · {x['dim']}燃油率：{_fmt_money(x['old'])} → {_fmt_money(x['new'])}  ({_pct_str(x['pct'])}) {arrow}")
            elif x["kind"] == "reg":
                L.append(f"   · {x['dim']} {x['w']}：{_fmt_money(x['old'])} → {_fmt_money(x['new'])} 元  ({_pct_str(x['pct'])}) {arrow}")
            else:
                L.append(f"   · {x['dim']} {x['w']}：{_fmt_money(x['old'])} → {_fmt_money(x['new'])} 元/kg  "
                         f"Δ{_fmt_money(x['delta'])} ({_pct_str(x['pct'])}) {arrow}")
    L.append("")
    L.append(f"📎 完整比价/历史看板：{SITE_URL}")
    return "\n".join(L)


# ---------- 环比通知（表格快照图，直观可直接转发） ----------
def _cjk_font(size, bold=False):
    import glob, os
    cands = []
    cands += glob.glob("/usr/share/fonts/**/NotoSansCJK*.*", recursive=True)
    cands += glob.glob("/usr/share/fonts/**/wqy*.*", recursive=True)
    cands += glob.glob("/usr/share/fonts/**/SourceHanSans*.*", recursive=True)
    cands += ["C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/msyhbd.ttc",
              "C:/Windows/Fonts/simhei.ttf", "C:/Windows/Fonts/simsun.ttc"]
    fp = os.environ.get("RB_FONT")
    if fp:
        cands = [fp] + cands
    for p in cands:
        if os.path.exists(p):
            try:
                from PIL import ImageFont
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    from PIL import ImageFont
    return ImageFont.load_default()


def render_channel_images(rep, prev_label, new_label):
    """按渠道逐张渲染环比快照图：每个「有价格环比变化」的渠道单独一张
    （含该渠道下所有国家 × 重量段的环比变化）；单渠道行数过多自动分页。
    返回所有图片路径列表（按渠道顺序，渠道内按页顺序）。无变化的渠道不会出现。"""
    from PIL import Image, ImageDraw
    if not rep:
        return []
    # 逐渠道快照不再画「渠道」列（表头已标明渠道），把空间留给国家/重量段
    # 重量段为完整符号区间（如「>0 ≤0.5kg」「>21 <45kg」「>21kg」），需更宽
    cols = [("国家/区域", 110), ("重量段", 232), ("类别", 80),
            ("旧值", 92), ("新值", 92), ("涨跌", 92), ("涨跌率", 92)]
    PAD = 14
    TOP = 120
    HEADER_H = 36
    ROW_H = 30
    W = PAD * 2 + sum(w for _, w in cols)

    def _fit(d, font, text, max_w):
        text = str(text).replace("\n", " ").replace("\r", " ")
        if d.textlength(text, font=font) <= max_w:
            return text
        for i in range(len(text) - 1, 0, -1):
            s = text[:i] + "..."
            if d.textlength(s, font=font) <= max_w:
                return s
        return "..."

    def fmt(x):
        if x["kind"] == "fuel":
            ov = "—" if x["old"] is None else f"{x['old']*100:.2f}%"
            nv = "—" if x["new"] is None else f"{x['new']*100:.2f}%"
            dv = "—" if x["delta"] is None else f"{x['delta']*100:+.2f}pp"
        else:
            ov = "—" if x["old"] is None else f"{x['old']:.2f}"
            nv = "—" if x["new"] is None else f"{x['new']:.2f}"
            dv = "—" if x["delta"] is None else f"{x['delta']:+.2f}"
        pv = "—" if x["pct"] is None else f"{x['pct']:+.1f}%"
        cat = {"rate": "运费/kg", "reg": "挂号费", "fuel": "燃油率"}.get(x["kind"], "")
        return ov, nv, dv, pv, cat

    COLORS = {"涨": (212, 56, 13), "跌": (56, 158, 13), "新增": (9, 109, 217),
              "移除": (140, 140, 140), "平": (51, 51, 51)}
    PER = 40
    paths = []
    page_no = 0
    for c, changes in rep:
        name = c.get("name", "") or c.get("carrier", "")
        if not changes:
            continue
        up = sum(1 for x in changes if x["direction"] == "涨")
        dn = sum(1 for x in changes if x["direction"] == "跌")
        newc = sum(1 for x in changes if x["direction"] == "新增")
        rmc = sum(1 for x in changes if x["direction"] == "移除")
        chunks = [changes[i:i + PER] for i in range(0, len(changes), PER)]
        for pi, chunk in enumerate(chunks):
            page_no += 1
            H = TOP + HEADER_H + len(chunk) * ROW_H + 24
            img = Image.new("RGB", (W, H), (255, 255, 255))
            d = ImageDraw.Draw(img)
            f_title = _cjk_font(19, True)
            d.text((PAD, 14), _fit(d, f_title, name, W - PAD * 2), fill=(20, 30, 50), font=f_title)
            f_sub = _cjk_font(12)
            d.text((PAD, 46), f"物流报价环比变动快照   |   环比基准：{prev_label}  ->  {new_label}",
                   fill=(90, 90, 90), font=f_sub)
            d.text((PAD, 68), f"涨价 {up}  降价 {dn}  新增 {newc}  移除 {rmc}  共 {len(changes)} 条变动",
                   fill=(90, 90, 90), font=f_sub)
            if len(chunks) > 1:
                d.text((W - PAD - 96, 68), f"第 {pi+1}/{len(chunks)} 页", fill=(140, 140, 140), font=f_sub)
            # 表头
            y = TOP
            d.rectangle([PAD, y, W - PAD, y + HEADER_H], fill=(31, 58, 95))
            cx = PAD
            f_h = _cjk_font(13, True)
            for title, w in cols:
                d.text((cx + 6, y + 10), title, fill=(255, 255, 255), font=f_h)
                cx += w
            # 数据行
            ry = y + HEADER_H
            f_c = _cjk_font(12)
            for idx, x in enumerate(chunk):
                if idx % 2 == 1:
                    d.rectangle([PAD, ry, W - PAD, ry + ROW_H], fill=(245, 247, 250))
                ov, nv, dv, pv, cat = fmt(x)
                color = COLORS.get(x["direction"], (51, 51, 51))
                cells = [x["dim"], x["w"], cat, ov, nv, dv, pv]
                cx = PAD
                for ci, (val, (_, w)) in enumerate(zip(cells, cols)):
                    fill = color if ci in (5, 6) else (51, 51, 51)
                    txt = _fit(d, f_c, str(val), w - 10)
                    tw = d.textlength(txt, font=f_c)
                    tx = cx + (w - tw) / 2
                    ty = ry + (ROW_H - 12) / 2
                    d.text((tx, ty), txt, fill=fill, font=f_c)
                    cx += w
                ry += ROW_H
            # 网格线
            cx = PAD
            for _, w in cols:
                d.line([cx, TOP, cx, ry], fill=(225, 228, 232))
                cx += w
            d.line([W - PAD, TOP, W - PAD, ry], fill=(225, 228, 232))
            d.line([PAD, TOP, W - PAD, TOP], fill=(225, 228, 232))
            d.line([PAD, ry, W - PAD, ry], fill=(225, 228, 232))
            d.line([PAD, TOP, PAD, ry], fill=(225, 228, 232))
            path = os.path.join(HERE, f"_ringbi_{uuid.uuid4().hex}.png")
            img.save(path)
            paths.append(path)
    return paths


def send_image(path):
    """上传 PNG 到飞书并以图片消息推送到群（可直接转发）。"""
    if not APP_SECRET:
        log("未配置 FEISHU_APP_SECRET，跳过发图")
        return
    try:
        t = req_json("https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal", "POST",
                     body={"app_id": APP_ID, "app_secret": APP_SECRET})
        tok = t["tenant_access_token"]
        import urllib.request as _ur
        boundary = "----rb" + str(int(time.time() * 1000))
        data = b""
        data += f"--{boundary}\r\n".encode()
        data += b'Content-Disposition: form-data; name="image_type"\r\n\r\n'
        data += b"message\r\n"
        data += f"--{boundary}\r\n".encode()
        data += b'Content-Disposition: form-data; name="image"; filename="ringbi.png"\r\n'
        data += b"Content-Type: image/png\r\n\r\n"
        data += open(path, "rb").read()
        data += f"\r\n--{boundary}--\r\n".encode()
        req = _ur.Request("https://open.feishu.cn/open-apis/im/v1/images", data=data,
                          headers={"Authorization": "Bearer " + tok,
                                   "Content-Type": f"multipart/form-data; boundary={boundary}"},
                          method="POST")
        with _ur.urlopen(req, timeout=60) as r:
            up = json.loads(r.read().decode())
        if up.get("code") != 0:
            log("图片上传失败: " + str(up)[:200])
            return
        key = up["data"]["image_key"]
        body = {"receive_id": CHAT_ID, "msg_type": "image", "content": json.dumps({"image_key": key})}
        r2 = req_json("https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id", "POST", tok, body)
        if r2.get("code") == 0:
            log("飞书图片发送成功 -> " + CHAT_ID)
        else:
            log("飞书图片消息错误: " + str(r2)[:200])
    except Exception as e:
        log(f"发图失败: {e}")


def send_text(text):
    """飞书文本消息（msg_type=text），可直接转发。超长自动分条。"""
    if not APP_SECRET:
        log("未配置 FEISHU_APP_SECRET，跳过发文本")
        return
    try:
        t = req_json("https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal", "POST",
                     body={"app_id": APP_ID, "app_secret": APP_SECRET})
        tok = t["tenant_access_token"]
        MAX = 28000
        chunks = [text[i:i+MAX] for i in range(0, len(text), MAX)] or [text]
        for i, ch in enumerate(chunks):
            body = {"receive_id": CHAT_ID, "msg_type": "text", "content": json.dumps({"text": ch})}
            r = req_json("https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id", "POST", tok, body)
            if r.get("code") == 0:
                log(f"飞书文本发送成功({i+1}/{len(chunks)}) -> " + CHAT_ID)
            else:
                log("飞书文本错误: " + str(r)[:200])
    except Exception as e:
        log(f"发文本失败: {e}")


def send_card(card):
    if not APP_SECRET:
        log("未配置 FEISHU_APP_SECRET，跳过发卡片")
        return
    try:
        t = req_json("https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal", "POST",
                     body={"app_id": APP_ID, "app_secret": APP_SECRET})
        tok = t["tenant_access_token"]
        body = {"receive_id": CHAT_ID, "msg_type": "interactive", "content": json.dumps(card)}
        r = req_json("https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=chat_id", "POST", tok, body)
        if r.get("code") == 0:
            log("飞书卡片发送成功 -> " + CHAT_ID)
        else:
            log("飞书卡片错误: " + str(r)[:200])
    except Exception as e:
        log(f"发卡片失败: {e}")


# ---------- bake rates.js ----------
def bake_ratesjs():
    import time as _t
    d = json.load(open(RATES, encoding="utf-8"))
    today = _t.strftime("%Y-%m-%d")
    now = _t.strftime("%Y-%m-%dT%H:%M:%S")
    d["generated"] = today              # 修复：显示系统同步日，不再写死 effective_date(8/04)
    meta = d.setdefault("meta", {})
    meta["baked_at"] = now              # 烘焙时间戳
    # 动态 banner：同步日 + 各家报价单真实主题（备注真实报价日，避免"显示同步日但报价是更早日"的误导）
    qs = dict(QUOTE_SOURCES) or dict(meta.get("quote_sources") or {})
    meta["quote_sources"] = qs
    parts = [f"{k}报价：{v}" for k, v in qs.items() if v]
    meta["banner"] = "数据同步于 " + today + (" ｜ " + "；".join(parts) if parts else "")
    eff = (d.get("meta") or {}).get("effective_date") or today   # 保留供下方 log 使用
    BANNER = "\n;(function(){function render(){var R=window.RATES||{};var g=R.generated||(R.meta&&R.meta.effective_date);var note=(R.meta&&R.meta.banner)||(R.meta&&R.meta.note)||'';var h=document.querySelector('header');if(!h||!g)return;var el=document.getElementById('effBanner');if(!el){el=document.createElement('div');el.id='effBanner';el.style.cssText='margin-top:8px;padding:6px 12px;border-radius:8px;font-size:13px;font-weight:600;color:#0e1117;background:linear-gradient(90deg,#ffd479,#ffb347);box-shadow:0 1px 4px rgba(0,0,0,.25);display:inline-block';h.appendChild(el);}el.textContent='\\U0001F4C5 '+note;}if(document.readyState!=='loading')render();else document.addEventListener('DOMContentLoaded',render);})();\n"
    with open(RATES_JS, "w", encoding="utf-8") as f:
        f.write("window.RATES = ")
        json.dump(d, f, ensure_ascii=False)
        f.write(";")
        f.write(BANNER)
    json.dump({"build": str(int(_t.time()))}, open(VERSION, "w", encoding="utf-8"))
    log(f"rates.js 重新烘焙 | effective_date={eff}")


def send_error_card(title, msg):
    send_card({
        "config": {"wide_screen_mode": True},
        "header": {"title": {"tag": "plain_text", "content": title}, "template": "red"},
        "elements": [{"tag": "div", "text": {"tag": "lark_md",
                  "content": "**自动抓单运行异常**（需人工排查）\n```\n" + msg[-1800:] + "\n```"}}],
    })


def main():
    try:
        _main_impl()
    except SystemExit:
        raise
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        log("❌ 运行异常: " + str(e))
        send_error_card("比价易 · 抓单异常", tb)
        raise


def _main_impl():
    if not os.path.exists(RATES):
        log("[skip] 无 rates.json")
        return
    old = json.load(open(RATES, encoding="utf-8"))
    token = load_token()
    if token is None and not FETCH_LOCAL:
        log("❌ 无法获取 wl02 邮箱 token，本次运行直接失败（不会静默跳过）")
        send_error_card("比价易 · 抓单失败", "无法获取 wl02 邮箱 user token：access_token 已过期且 refresh_token 续期失败。\n请在本机重跑 quote_pull/oauth_authorize.py 让用户在浏览器点一次「允许」，重新写入 WL02_MAIL_TOKEN secret。")
        sys.exit(1)
    # 抓取各承运商报价单
    got_any = False
    for name, cfg in CARRIERS.items():
        try:
            if process_carrier(token, name, cfg):
                got_any = True
        except Exception as e:
            log(f"[{name}] 处理异常: {e}")
    # 调 bake 解析（只改价格数值）
    for name, scripts in BAKE.items():
        for sc in scripts:
            try:
                subprocess.run([sys.executable, os.path.join(HERE, sc), "--write"],
                               check=True, capture_output=True, text=True, timeout=240)
                log(f"[{name}] {sc} 完成")
            except subprocess.CalledProcessError as e:
                log(f"[{name}] {sc} 失败: {e.stderr[-300:] if e.stderr else e}")
    new = json.load(open(RATES, encoding="utf-8"))
    # 白名单守铁律
    if not whitelist_ok(old, new):
        log("❌ 白名单 FAIL：结构/渠道被改动，禁止提交！已恢复原 rates.json")
        json.dump(old, open(RATES, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        return
    # 每次抓单都重新烘焙 rates.js：刷新 generated(同步日)/baked_at/banner（即使价格无变化也更新日期）
    bake_ratesjs()
    # 计算变更（白名单已确保只改数值；有变化才继续）
    added, removed, changed = compute_diff(old, new)
    if not (added or removed or changed):
        log("✅ 无价格数值变化，跳过提交与通知")
        if os.environ.get("FETCH_DIAG") == "1":
            dlog = ["**【诊断模式】本次自动抓单明细**", ""]
            dlog += REPORT or ["（无各承运商处理记录）"]
            dlog += ["", "结论：邮箱抓取/解析均执行完毕，当前 rates.json 与最新报价单相比无价格数值变化。"]
            send_card({
                "config": {"wide_screen_mode": True},
                "header": {"title": {"tag": "plain_text", "content": "比价易 · 抓单诊断"}, "template": "grey"},
                "elements": [{"tag": "div",
                              "text": {"tag": "lark_md", "content": "\n".join(dlog)}}],
            })
        return
    # 计算价格环比（仅运费/kg / 挂号费 / 燃油率 三类价格数值）
    ringbi = compute_ringbi(old, new)
    # （rates.js 已在上方抓单阶段统一重新烘焙，此处不再重复）
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    _raw_prev = (old.get("meta") or {}).get("baked_at") or (old.get("meta") or {}).get("effective_date") or "上一版"
    prev_label = _raw_prev[:10] if isinstance(_raw_prev, str) and len(_raw_prev) >= 10 else _raw_prev
    # 只推「价格环比有变化」的快照图；无任何价格环比变化则不往群里推消息
    if ringbi:
        imgs = render_channel_images(ringbi, prev_label, "本版(" + now + ")")
        for p in imgs:
            send_image(p)
            try:
                os.remove(p)
            except Exception:
                pass
        log(f"✅ 价格环比变动 {len(ringbi)} 个渠道 / {sum(len(c) for _, c in ringbi)} 项，"
            f"已按渠道逐张推送 {len(imgs)} 张快照图，待 yml 提交")
    else:
        log("✅ 有数值变化但非运费价格环比，已更新价格未推图")


if __name__ == "__main__":
    main()
