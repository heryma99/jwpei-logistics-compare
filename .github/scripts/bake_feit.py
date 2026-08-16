# -*- coding: utf-8 -*-
# Bake 飞特(Feit) 小包渠道 from 飞特/latest.xlsx (价格表，如「75571 飞特标准挂号普货（EY）价格表 7-9.xlsx」).
# 仅更新已存在渠道 feit_ca_ey 的 brackets（加拿大全境同价）；保留结构(18档)与元数据。
# Dry-run by default; --write to apply into rates.json.
import openpyxl, json, re, os, argparse

HERE = os.path.dirname(os.path.abspath(__file__))
NEWXLSX = os.path.join(HERE, "飞特", "latest.xlsx")
RATES = os.path.join(HERE, "..", "..", "rates.json")
OUT = os.path.join(HERE, "patch_feit.json")

CHANNEL_ID = "feit_ca_ey"
SHEET = "飞特标准挂号普货(EY)"
COUNTRY = "加拿大"

def num(x):
    try:
        if x is None:
            return None
        s = str(x).replace("，", "").replace(",", "").strip()
        if s in ("", "***", "/", "-", "—"):
            return None
        return float(s)
    except Exception:
        return None

def extract_brackets():
    """从价格表 sheet 解析 [{up, rate, reg}, ...]；C列重量文本取上限，E列KG/RMB，F列Item/RMB。"""
    wb = openpyxl.load_workbook(NEWXLSX, read_only=True, data_only=True)
    ws = wb[SHEET]
    # 定位表头行（含「渠道代码」）
    hrow = None
    for ri in range(1, 14):
        row = [c for c in next(ws.iter_rows(min_row=ri, max_row=ri, max_col=10, values_only=True))]
        for v in row:
            if v is not None and str(v).startswith("渠道代码"):
                hrow = ri
                break
        if hrow:
            break
    if hrow is None:
        wb.close()
        raise RuntimeError("未找到表头(渠道代码)")
    hdr = [c for c in next(ws.iter_rows(min_row=hrow, max_row=hrow, max_col=10, values_only=True))]
    widx = ridx = gidx = None
    for i, v in enumerate(hdr):
        s = str(v) if v is not None else ""
        # 注意「最低计费重量（KG）」也含"重量+KG"，必须排除，否则 widx 会被覆盖成最低计费重量列
        if "重量" in s and "KG" in s and "最低" not in s:
            widx = i
        elif "KG/RMB" in s or ("运费" in s and "KG" in s):
            ridx = i
        elif "Item/RMB" in s or "挂号" in s:
            gidx = i
    if None in (widx, ridx, gidx):
        wb.close()
        raise RuntimeError(f"列定位失败 widx={widx} ridx={ridx} gidx={gidx}")
    brackets = []
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        cells = list(row)
        if len(cells) <= max(widx, ridx, gidx):
            continue
        wtext = cells[widx]
        rate = num(cells[ridx])
        reg = num(cells[gidx])
        if wtext is None or rate is None:
            continue
        nums = re.findall(r"[\d.]+", str(wtext))
        if not nums:
            continue
        try:
            up = float(nums[-1])
            reg = reg if reg is not None else 0.0
        except Exception:
            continue
        brackets.append({"up": up, "rate": rate, "reg": reg})
    wb.close()
    return brackets

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()
    if not os.path.exists(NEWXLSX):
        print(f"[跳过] 无 {NEWXLSX}（尚未抓到报价单）")
        return
    rates = json.load(open(RATES, encoding="utf-8"))
    byid = {c["id"]: c for c in rates["channels"]}
    ch = byid.get(CHANNEL_ID)
    if not ch:
        print(f"[跳过] {CHANNEL_ID}: rates.json 无此渠道")
        return
    try:
        new = extract_brackets()
    except Exception as e:
        print(f"[错误] {CHANNEL_ID}: {e}")
        return
    old_block = ch.get("countries", {}).get(COUNTRY, {})
    old = old_block.get("brackets", [])
    extra = {k: v for k, v in old_block.items() if k != "brackets"}
    new_country = {**extra, "brackets": new}
    full = {k: v for k, v in ch.get("countries", {}).items()}
    full[COUNTRY] = new_country
    report = [f"[更新] {CHANNEL_ID} <- {SHEET}: 旧档数 {len(old)} -> 新档数 {len(new)}"]
    if old and new:
        report.append(f"    首档 {old[0]} -> {new[0]}")
        report.append(f"    末档 {old[-1]} -> {new[-1]}")
    print("\n".join(report))
    patch = {CHANNEL_ID: {"countries": full, "dest": list(full.keys())}}
    json.dump(patch, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\nWROTE {OUT}")
    if not args.write:
        print("[DRY-RUN] re-run with --write to apply")
        return
    ch["countries"] = full
    ch["dest"] = list(full.keys())
    json.dump(rates, open(RATES, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\n[WRITE] applied {CHANNEL_ID}")

if __name__ == "__main__":
    main()
